import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── Helpers ────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)

GREEN_FILL = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid')
RED_FILL = PatternFill(start_color='ef4444', end_color='ef4444', fill_type='solid')


def _score_fill(score):
    if score >= 80:
        return GREEN_FILL
    elif score >= 50:
        return ORANGE_FILL
    return RED_FILL


def _style_header_row(ws, row_num):
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


# ── 1. Excel report ────────────────────────────────────────────────────────────

def generate_excel_report(df_clean: pd.DataFrame, results: dict) -> bytes:
    wb = Workbook()

    # Sheet 1 — Cleaned data
    ws1 = wb.active
    ws1.title = 'Données nettoyées'
    ws1.append(list(df_clean.columns))
    _style_header_row(ws1, 1)
    for row in df_clean.itertuples(index=False):
        ws1.append(list(row))
    _auto_width(ws1)

    # Sheet 2 — Quality report
    ws2 = wb.create_sheet('Rapport qualité')
    headers2 = [
        'Colonne', 'Type détecté', 'Score avant (%)', 'Score après (%)', 'Amélioration',
        'Complétude (%)', 'Validité (%)', 'Unicité (%)',
        'Valeurs manquantes', 'Valeurs invalides', 'Doublons',
    ]
    ws2.append(headers2)
    _style_header_row(ws2, 1)

    col_types = results.get('col_types', {})
    scores_before = results.get('col_scores_before', {})
    scores_after = results.get('col_scores_after', {})

    for col in df_clean.columns:
        sb = scores_before.get(col, {})
        sa = scores_after.get(col, {})
        score_b = sb.get('score', 0)
        score_a = sa.get('score', 0)
        row = [
            col,
            col_types.get(col, 'text'),
            score_b,
            score_a,
            round(score_a - score_b, 1),
            sa.get('completeness', 0),
            sa.get('validity', 0),
            sa.get('uniqueness', 0),
            sa.get('missing_count', 0),
            sa.get('invalid_count', 0),
            sa.get('duplicate_count', 0),
        ]
        ws2.append(row)
        last_row = ws2.max_row
        # Color the "Score après" cell (column D = index 4)
        score_cell = ws2.cell(row=last_row, column=4)
        score_cell.fill = _score_fill(score_a)
        score_cell.font = Font(color='FFFFFF', bold=True)

    _auto_width(ws2)

    # Sheet 3 — Summary
    ws3 = wb.create_sheet('Résumé')
    ws3.append(['Métrique', 'Valeur'])
    _style_header_row(ws3, 1)
    summary_rows = [
        ('Date de génération', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Lignes avant nettoyage', results.get('rows_before', 0)),
        ('Lignes après nettoyage', results.get('rows_after', 0)),
        ('Colonnes traitées', len(df_clean.columns)),
        ('Score global avant (%)', results.get('global_score_before', 0)),
        ('Score global après (%)', results.get('global_score_after', 0)),
        ('Amélioration globale', round(
            results.get('global_score_after', 0) - results.get('global_score_before', 0), 1
        )),
    ]
    for r in summary_rows:
        ws3.append(list(r))
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 2. CSV report ──────────────────────────────────────────────────────────────

def generate_csv_report(df_clean: pd.DataFrame) -> bytes:
    return df_clean.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')


# ── 3. JSON summary ────────────────────────────────────────────────────────────

def build_json_summary(results: dict) -> dict:
    col_types = results.get('col_types', {})
    scores_before = results.get('col_scores_before', {})
    scores_after = results.get('col_scores_after', {})

    columns = {}
    for col in col_types:
        columns[col] = {
            'type': col_types[col],
            'score_before': scores_before.get(col, {}).get('score', 0),
            'score_after': scores_after.get(col, {}).get('score', 0),
        }

    return {
        'generated_at': datetime.now().isoformat(),
        'rows_before': results.get('rows_before', 0),
        'rows_after': results.get('rows_after', 0),
        'global_score_before': results.get('global_score_before', 0),
        'global_score_after': results.get('global_score_after', 0),
        'columns': columns,
        'changes': results.get('changes_log', []),
    }
